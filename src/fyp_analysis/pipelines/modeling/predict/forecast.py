import shutil
from datetime import datetime

import pandas as pd
from fbprophet import Prophet
from openpyxl import load_workbook

from fyp_analysis.pipelines.data_processing import preprocess

from .... import SRC_DIR
from ....extras.datasets import load_cbo_data
from ...data_processing.preprocess import PreprocessPipeline
from .core import get_forecasts_from_fits
from .utils import aggregate_to_fiscal_year


def _project_flat_growth(X, start="2019-04", stop="2020-03"):
    """Normalize future growth to be flat at the last annual period."""
    # Make a copy first
    X = X.copy()
    freq = X.index.inferred_freq
    X.index.freq = freq

    # This is the part that will be projected
    norm = X.loc[start:stop].copy()
    latest_date = norm.index[-1]

    # This should be monthly or quarterly
    assert len(norm) in [4, 12]
    if len(norm) == 4:
        key = lambda dt: dt.quarter
    else:
        key = lambda dt: dt.month

    # Reset the index to months/quarters
    norm.index = [key(dt) for dt in norm.index]

    # Change the forecast to be flat
    forecast_start = latest_date + latest_date.freq
    Y = X.loc[forecast_start:].copy()

    # Reset index
    i = Y.index
    Y.index = [key(dt) for dt in Y.index]

    # Overwrite
    Y.loc[:] = norm.loc[Y.index].values
    Y.index = i

    # Add back to original
    X.loc[Y.index] = Y.values

    return X


def get_var_forecast(
    unscaled_features,
    stationary_guide,
    fit_params,
    tax_base_name,
    plan_start_year,
    cbo_forecast_date,
):

    # Initialize the preprocess pipeline
    preprocess = PreprocessPipeline(stationary_guide)

    # Load the cbo data
    cbo_data = load_cbo_data(date=cbo_forecast_date)

    # Get all of the fits
    all_forecasts = get_forecasts_from_fits(
        unscaled_features,
        preprocess,
        fit_params,
        tax_base_name,
        plan_start_year,
        cbo_data,
        max_fits=len(fit_params),
    )

    # Take the average
    average_fit = all_forecasts.mean(axis=1).squeeze().rename(tax_base_name)

    return average_fit


def get_prophet_forecast(
    future_ratios, unscaled_features, tax_base_column, plan_start_year
):

    df = (
        unscaled_features[tax_base_column]
        .rename_axis("ds")
        .reset_index()
        .rename(columns={tax_base_column: "y"})
        .sort_values("ds")
    )

    fit_kwargs = {
        "daily_seasonality": False,
        "weekly_seasonality": False,
        "n_changepoints": min(int(0.7 * len(df)), 25),
        "seasonality_mode": "additive",
    }

    # Fit the model on the pre-covid data
    precovid = df.query("ds < '2020-04-01' and ds >= '2014-01-01'")
    model = Prophet(**fit_kwargs)
    model.fit(precovid)

    # Get the forecast period
    freq = "Q"
    forecast_stop_date = f"{plan_start_year+4}-06-30"
    periods = (
        pd.to_datetime(forecast_stop_date).to_period(freq)
        - df["ds"].max().to_period(freq)
    ).n
    future = model.make_future_dataframe(periods=periods + 5, freq="QS")

    # Forecast
    forecast = model.predict(future)

    # Format
    forecast = (
        forecast[
            [
                "ds",
                "yhat",
                "yhat_lower",
                "yhat_upper",
                "yearly",
                "trend",
            ]
        ]
        .rename(
            columns={
                "ds": "date",
                "yhat": "total",
                "yhat_lower": "lower",
                "yhat_upper": "upper",
            }
        )
        .set_index("date")
    )

    # Get the flat forecast
    flat_forecast = _project_flat_growth(forecast)

    # Combine the forecast and actuals
    F = flat_forecast["total"].copy()
    actuals = df.set_index("ds")["y"]
    F = pd.concat([F.loc["2021-07-01":], actuals])

    # Apply the ratios
    inter = F.index.intersection(future_ratios.index)
    F.loc[inter] *= future_ratios.loc[inter].values

    return aggregate_to_fiscal_year(F).squeeze().rename(tax_base_column)


def report_forecast_results(plan_start_year, tax_revenues, tax_bases):
    """Report the forecast results using a template spreadsheet."""

    # Copy over the template
    template_path = SRC_DIR / "templates" / "revenue_summary_template.xlsx"
    filename = SRC_DIR / ".." / ".." / "data" / "07_reporting" / "revenue_summary.xlsx"
    shutil.copy(template_path, filename)

    # Load the file
    book = load_workbook(filename)

    # Verify sheets are deleted
    sheets = [
        "Revenue Data",
        "Tax Base Data",
    ]
    for sheet in sheets:
        if sheet in book.sheetnames:
            del book[sheet]
    book.save(filename)

    # Open and save the calculations
    with pd.ExcelWriter(filename, engine="openpyxl", mode="a") as writer:

        # Trim by fiscal year
        X = tax_revenues.query(f"fiscal_year >= {plan_start_year-1}")
        Y = tax_bases.query(f"fiscal_year >= {plan_start_year-1}")

        # Reorder rows and columns
        X = (
            X.set_index("tax_name")
            .loc[["Wage", "Sales", "BIRT", "RTT", "Parking", "Amusement", "NPT"]]
            .reset_index()[["fiscal_year", "Five Year Plan", "Controller", "tax_name"]]
        )

        # Save revenue data
        X.to_excel(writer, sheet_name="Revenue Data", index=False)

        # Re-order the columns
        Y = Y[
            [
                "WageBase",
                "SalesBase",
                "BIRTBase",
                "GrossReceiptsBase",
                "NetIncomeBase",
                "RTTBase",
                "ParkingBase",
                "AmusementBase",
                "NPTBase",
            ]
        ].reset_index()

        # Save tax base data
        Y.to_excel(writer, sheet_name="Tax Base Data", index=False)

    # Update dates
    book = load_workbook(filename)

    # Update the headers
    sheets = ["Revenue by FY", "Revenue by Tax"]
    for sheet_name in sheets:

        ws = book[sheet_name]
        ws["A2"] = f"The Five Year Plan: FY {plan_start_year} - FY {plan_start_year+4}"
        ws["A4"] = datetime.today().strftime("%m/%d/%Y")

    # --------------------------------
    # Section headers on FIRST SHEET
    # --------------------------------
    row = 6
    ws = book["Revenue by FY"]
    for yr in range(plan_start_year, plan_start_year + 5):
        ws[f"A{row}"] = f"FY {yr}"
        row += 12

    # Do the summary too
    ws[f"A{row}"] = f"FY {plan_start_year} - FY {plan_start_year+4}"

    # --------------------------------
    # Section headers on SECOND SHEET
    # --------------------------------
    start = 7
    ws = book["Revenue by Tax"]
    for n in range(9):
        row = start + 8 * n
        for i, yr in enumerate(range(plan_start_year, plan_start_year + 5)):
            ws[f"A{row + i}"] = f"FY {yr}"

    # --------------------------------
    # Section headers on THIRD SHEET
    # --------------------------------
    start = 9
    ws = book["Growth by Tax"]
    for n in range(7):
        row = start + 7 * n
        for i, yr in enumerate(range(plan_start_year, plan_start_year + 5)):
            ws[f"B{row + i}"] = f"FY {yr}"

    # --------------------------------
    # Section headers on FOURTH SHEET
    # --------------------------------
    row = 7
    ws = book["Growth by Year"]
    for yr in range(plan_start_year, plan_start_year + 5):
        ws[f"B{row}"] = f"FY {yr}"
        row += 9

    # Update the headers
    sheets = ["Growth by Tax", "Growth by Year"]
    for sheet_name in sheets:

        ws = book[sheet_name]
        ws["B2"] = f"The Five Year Plan: FY {plan_start_year} - FY {plan_start_year+4}"

    # Save the book
    book.save(filename)
